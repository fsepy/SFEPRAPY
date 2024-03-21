from typing import Dict

import openpyxl


def xlsx_to_dict(fp: str) -> Dict[str, dict]:
    # Load the workbook
    wb = openpyxl.load_workbook(fp)

    # Select the first worksheet
    sheet = wb.active

    # Get the headers (first row) and the row headers (first column)
    headers = [cell.value for cell in sheet[1]]
    headers = headers[1:]
    row_headers = [cell.value for cell in sheet['A'][1:]]

    # Iterate through the sheet and create the nested dictionary
    result = {}
    for header, col_idx in zip(headers, range(2, len(headers) + 2)):
        col_values = {}
        for row_header, row_idx in zip(row_headers, range(2, len(row_headers) + 2)):
            col_values[row_header] = sheet.cell(row=row_idx, column=col_idx).value
        result[header] = col_values

    return result


def dict_to_xlsx(data: Dict[str, dict], fp: str):
    """

    :param data:
    :param fp:
    :return:

    data = {
        "Column1": {"Row1": "A1", "Row2": "A2", "Row3": "A3"},
        "Column2": {"Row1": "B1", "Row2": "B2", "Row3": "B3"},
        "Column3": {"Row1": "C1", "Row2": "C2", "Row3": "C3"},
    }
    """
    # check all nested dict
    row_headers = None
    for k, v in data.items():
        if row_headers:
            if row_headers != v.keys():
                missing_keys = list()
                missing_keys.extend([_ for _ in row_headers if _ not in v.keys()])
                missing_keys.extend([_ for _ in v.keys() if _ not in row_headers])
                raise IndexError(f'keys do not match between dicts in the list: {missing_keys}')
        row_headers = v.keys()
    row_headers = tuple(row_headers)

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write the column headers
    for col, column_name in enumerate(data.keys(), start=1):
        ws.cell(row=1, column=col + 1, value=column_name)

    # Write the row headers
    for row, row_name in enumerate(row_headers):
        ws.cell(row=row + 2, column=1, value=row_name)

    # Write the data to the worksheet
    for col, (column_name, data_) in enumerate(data.items(), start=1):
        for row, (row_name, value) in enumerate(data_.items(), start=2):
            ws.cell(row=row_headers.index(row_name) + 2, column=col + 1, value=value)

    # Save the workbook to an XLSX file
    wb.save(fp)
